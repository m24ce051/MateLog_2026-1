# tracking/admin_views.py
# Vistas personalizadas para el admin de tracking con visualización matricial

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.contrib.auth import get_user_model
from lessons.models import Tema
from .models import TiempoPantalla, ClicBoton
import csv

User = get_user_model()


@staff_member_required
def resumen_tiempos_view(request):
    """
    Vista matricial de tiempo por pantalla.
    Muestra: Usuario | Teoría 1 | Teoría 2 | Ejemplo 1 | ... | Total
    """
    # Obtener tema seleccionado del filtro
    tema_id = request.GET.get('tema')
    exportar = request.GET.get('exportar')

    # Obtener todos los temas para el dropdown
    temas = Tema.objects.all().order_by('leccion__orden', 'orden')

    # Construir query base
    tiempos_query = TiempoPantalla.objects.all()

    if tema_id:
        tiempos_query = tiempos_query.filter(tema_id=tema_id)
        tema_seleccionado = Tema.objects.get(id=tema_id)
    else:
        tema_seleccionado = None

    # Obtener estructura de contenidos del tema seleccionado
    contenidos_estructura = []
    if tema_seleccionado:
        # Obtener todos los registros del tema para saber qué contenidos existen
        contenidos_distintos = (
            tiempos_query
            .values('tipo_contenido', 'numero')
            .distinct()
            .order_by('tipo_contenido', 'numero')
        )

        for item in contenidos_distintos:
            tipo = item['tipo_contenido']
            numero = item['numero']
            tipo_display = dict(TiempoPantalla.TIPO_CONTENIDO_CHOICES)[tipo]
            contenidos_estructura.append({
                'tipo': tipo,
                'numero': numero,
                'nombre': f"{tipo_display} {numero}"
            })

    # Obtener datos agregados por usuario
    datos_usuarios = []
    usuarios = User.objects.filter(tiempos_pantalla__in=tiempos_query).distinct()

    for usuario in usuarios:
        fila = {
            'usuario': usuario.username,
            'contenidos': {},
            'contenidos_lista': [],  # Lista ordenada para el template
            'total': 0
        }

        # Obtener tiempo por cada tipo de contenido
        tiempos_usuario = (
            tiempos_query
            .filter(usuario=usuario)
            .values('tipo_contenido', 'numero')
            .annotate(tiempo_total=Sum('tiempo_segundos'))
        )

        for tiempo in tiempos_usuario:
            key = f"{tiempo['tipo_contenido']}_{tiempo['numero']}"
            segundos = tiempo['tiempo_total']
            fila['contenidos'][key] = segundos
            fila['total'] += segundos

        # Crear lista ordenada de contenidos para el template
        for contenido in contenidos_estructura:
            key = f"{contenido['tipo']}_{contenido['numero']}"
            fila['contenidos_lista'].append(fila['contenidos'].get(key, 0))

        datos_usuarios.append(fila)

    # Calcular totales por columna
    totales_columnas = {}
    totales_lista = []  # Lista ordenada para el template
    total_general = 0

    for contenido in contenidos_estructura:
        key = f"{contenido['tipo']}_{contenido['numero']}"
        total_col = sum(
            usuario['contenidos'].get(key, 0)
            for usuario in datos_usuarios
        )
        totales_columnas[key] = total_col
        totales_lista.append(total_col)
        total_general += total_col

    # Manejar exportación
    if exportar == 'csv':
        return exportar_tiempos_csv(
            datos_usuarios, contenidos_estructura, totales_columnas,
            total_general, tema_seleccionado
        )
    elif exportar == 'excel':
        return exportar_tiempos_excel(
            datos_usuarios, contenidos_estructura, totales_columnas,
            total_general, tema_seleccionado
        )

    context = {
        'title': 'Resumen de Tiempos por Pantalla',
        'temas': temas,
        'tema_seleccionado': tema_seleccionado,
        'contenidos_estructura': contenidos_estructura,
        'datos_usuarios': datos_usuarios,
        'totales_columnas': totales_columnas,
        'totales_lista': totales_lista,
        'total_general': total_general,
    }

    return render(request, 'admin/tracking/resumen_tiempos.html', context)


@staff_member_required
def resumen_clics_view(request):
    """
    Vista matricial de clics por botón.
    Muestra: Usuario | Regresar | Ir a ejercicios | Volver | ... | Total
    """
    # Obtener tema seleccionado del filtro
    tema_id = request.GET.get('tema')
    exportar = request.GET.get('exportar')

    # Obtener todos los temas para el dropdown
    temas = Tema.objects.all().order_by('leccion__orden', 'orden')

    # Construir query base
    clics_query = ClicBoton.objects.all()

    if tema_id:
        clics_query = clics_query.filter(tema_id=tema_id)
        tema_seleccionado = Tema.objects.get(id=tema_id)
    else:
        tema_seleccionado = None

    # Tipos de botones (siempre los mismos 5)
    tipos_botones = ClicBoton.TIPO_BOTON_CHOICES

    # Obtener datos agregados por usuario
    datos_usuarios = []
    usuarios = User.objects.filter(clics_botones__in=clics_query).distinct()

    for usuario in usuarios:
        fila = {
            'usuario': usuario.username,
            'botones': {},
            'botones_lista': [],  # Lista ordenada para el template
            'total': 0
        }

        # Obtener clics por cada tipo de botón
        clics_usuario = (
            clics_query
            .filter(usuario=usuario)
            .values('tipo_boton')
            .annotate(cantidad=Count('id'))
        )

        for clic in clics_usuario:
            tipo = clic['tipo_boton']
            cantidad = clic['cantidad']
            fila['botones'][tipo] = cantidad
            fila['total'] += cantidad

        # Crear lista ordenada de botones para el template
        for tipo_key, tipo_display in tipos_botones:
            fila['botones_lista'].append(fila['botones'].get(tipo_key, 0))

        datos_usuarios.append(fila)

    # Calcular totales por columna
    totales_columnas = {}
    totales_lista = []  # Lista ordenada para el template
    total_general = 0

    for tipo_key, tipo_display in tipos_botones:
        total_col = sum(
            usuario['botones'].get(tipo_key, 0)
            for usuario in datos_usuarios
        )
        totales_columnas[tipo_key] = total_col
        totales_lista.append(total_col)
        total_general += total_col

    # Manejar exportación
    if exportar == 'csv':
        return exportar_clics_csv(
            datos_usuarios, tipos_botones, totales_columnas,
            total_general, tema_seleccionado
        )
    elif exportar == 'excel':
        return exportar_clics_excel(
            datos_usuarios, tipos_botones, totales_columnas,
            total_general, tema_seleccionado
        )

    context = {
        'title': 'Resumen de Clics por Botón',
        'temas': temas,
        'tema_seleccionado': tema_seleccionado,
        'tipos_botones': tipos_botones,
        'datos_usuarios': datos_usuarios,
        'totales_columnas': totales_columnas,
        'totales_lista': totales_lista,
        'total_general': total_general,
    }

    return render(request, 'admin/tracking/resumen_clics.html', context)


# ==================== FUNCIONES DE EXPORTACIÓN ====================

def exportar_tiempos_csv(datos_usuarios, contenidos_estructura, totales_columnas, total_general, tema):
    """Exporta la matriz de tiempos a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    tema_nombre = tema.titulo if tema else "Todos los temas"
    filename = f'tiempos_pantalla_{tema_nombre.replace(" ", "_")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Agregar BOM para Excel UTF-8
    response.write('\ufeff')

    writer = csv.writer(response)

    # Header
    header = ['Usuario']
    for contenido in contenidos_estructura:
        header.append(contenido['nombre'])
    header.append('Total')
    writer.writerow(header)

    # Datos de usuarios
    for usuario in datos_usuarios:
        fila = [usuario['usuario']]
        for contenido in contenidos_estructura:
            key = f"{contenido['tipo']}_{contenido['numero']}"
            fila.append(usuario['contenidos'].get(key, 0))
        fila.append(usuario['total'])
        writer.writerow(fila)

    # Fila de totales
    fila_totales = ['TOTAL']
    for contenido in contenidos_estructura:
        key = f"{contenido['tipo']}_{contenido['numero']}"
        fila_totales.append(totales_columnas.get(key, 0))
    fila_totales.append(total_general)
    writer.writerow(fila_totales)

    return response


def exportar_clics_csv(datos_usuarios, tipos_botones, totales_columnas, total_general, tema):
    """Exporta la matriz de clics a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    tema_nombre = tema.titulo if tema else "Todos los temas"
    filename = f'clics_botones_{tema_nombre.replace(" ", "_")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Agregar BOM para Excel UTF-8
    response.write('\ufeff')

    writer = csv.writer(response)

    # Header
    header = ['Usuario']
    for tipo_key, tipo_display in tipos_botones:
        header.append(tipo_display)
    header.append('Total')
    writer.writerow(header)

    # Datos de usuarios
    for usuario in datos_usuarios:
        fila = [usuario['usuario']]
        for tipo_key, tipo_display in tipos_botones:
            fila.append(usuario['botones'].get(tipo_key, 0))
        fila.append(usuario['total'])
        writer.writerow(fila)

    # Fila de totales
    fila_totales = ['TOTAL']
    for tipo_key, tipo_display in tipos_botones:
        fila_totales.append(totales_columnas.get(tipo_key, 0))
    fila_totales.append(total_general)
    writer.writerow(fila_totales)

    return response


def exportar_tiempos_excel(datos_usuarios, contenidos_estructura, totales_columnas, total_general, tema):
    """Exporta la matriz de tiempos a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Si no está instalado openpyxl, usar CSV como fallback
        return exportar_tiempos_csv(datos_usuarios, contenidos_estructura, totales_columnas, total_general, tema)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tiempos por Pantalla"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True)

    # Header
    headers = ['Usuario']
    for contenido in contenidos_estructura:
        headers.append(contenido['nombre'])
    headers.append('Total')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Datos de usuarios
    row_num = 2
    for usuario in datos_usuarios:
        ws.cell(row=row_num, column=1, value=usuario['usuario'])
        col_num = 2
        for contenido in contenidos_estructura:
            key = f"{contenido['tipo']}_{contenido['numero']}"
            ws.cell(row=row_num, column=col_num, value=usuario['contenidos'].get(key, 0))
            col_num += 1
        ws.cell(row=row_num, column=col_num, value=usuario['total'])
        row_num += 1

    # Fila de totales
    ws.cell(row=row_num, column=1, value='TOTAL').font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    col_num = 2
    for contenido in contenidos_estructura:
        key = f"{contenido['tipo']}_{contenido['numero']}"
        cell = ws.cell(row=row_num, column=col_num, value=totales_columnas.get(key, 0))
        cell.font = total_font
        cell.fill = total_fill
        col_num += 1
    cell = ws.cell(row=row_num, column=col_num, value=total_general)
    cell.font = total_font
    cell.fill = total_fill

    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Guardar a response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    tema_nombre = tema.titulo if tema else "Todos los temas"
    filename = f'tiempos_pantalla_{tema_nombre.replace(" ", "_")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def exportar_clics_excel(datos_usuarios, tipos_botones, totales_columnas, total_general, tema):
    """Exporta la matriz de clics a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Si no está instalado openpyxl, usar CSV como fallback
        return exportar_clics_csv(datos_usuarios, tipos_botones, totales_columnas, total_general, tema)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clics por Botón"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True)

    # Header
    headers = ['Usuario']
    for tipo_key, tipo_display in tipos_botones:
        headers.append(tipo_display)
    headers.append('Total')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Datos de usuarios
    row_num = 2
    for usuario in datos_usuarios:
        ws.cell(row=row_num, column=1, value=usuario['usuario'])
        col_num = 2
        for tipo_key, tipo_display in tipos_botones:
            ws.cell(row=row_num, column=col_num, value=usuario['botones'].get(tipo_key, 0))
            col_num += 1
        ws.cell(row=row_num, column=col_num, value=usuario['total'])
        row_num += 1

    # Fila de totales
    ws.cell(row=row_num, column=1, value='TOTAL').font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    col_num = 2
    for tipo_key, tipo_display in tipos_botones:
        cell = ws.cell(row=row_num, column=col_num, value=totales_columnas.get(tipo_key, 0))
        cell.font = total_font
        cell.fill = total_fill
        col_num += 1
    cell = ws.cell(row=row_num, column=col_num, value=total_general)
    cell.font = total_font
    cell.fill = total_fill

    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Guardar a response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    tema_nombre = tema.titulo if tema else "Todos los temas"
    filename = f'clics_botones_{tema_nombre.replace(" ", "_")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
